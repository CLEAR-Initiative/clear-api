#!/usr/bin/env python3
"""
MSNA Admin2 needs severity scoring script.

Reads IOM_MSNA_Admin2_Output_Tables.xlsx, applies per-sector scoring formulas,
optionally matches localities to DB pCodes, and writes a JSON output file.

Usage:
    python import_msna.py              # full run (requires DB)
    python import_msna.py --verify     # print column verification table and exit
    python import_msna.py --no-db      # skip pCode lookup (fully offline)

Source: IOM MSNA Admin2 Output Tables (August 2025)
Scoring: https://www.notion.so/Severity-Scoring-Logic-3585043895cf807a9214d0e16d0c9c2e
"""

import argparse
import json
import re
import sys
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "IOM_MSNA_Admin2_Output_Tables.xlsx"
OUTPUT_DIR = SCRIPT_DIR / "outputs"
CLEAR_API_URL = "https://dev-api.clearinitiative.io/graphql"

FORMULA_SOURCE = "https://www.notion.so/Severity-Scoring-Logic-3585043895cf807a9214d0e16d0c9c2e"

# ─── Formula documentation ────────────────────────────────────────────────────
# Included verbatim in the output metadata for traceability.
# sheet/col references map to IOM_MSNA_Admin2_Output_Tables.xlsx (0-indexed columns).

FORMULA_DOCS: dict = {
    "FSL": {
        "formula": "((poor_FCS * 0.70 + borderline_FCS * 0.30) + (rcsi_phase3 + rcsi_phase4) + (lcsi_crisis + lcsi_emergency)) / 3",
        "standard": "IPC Technical Manual v3.1 (https://www.ipcinfo.org/fileadmin/user_upload/ipcinfo/manual/IPC_Technical_Manual_3_Final.pdf) - three-pillar convergence of evidence: food consumption (FCS), coping strategies (rCSI), livelihoods (LCSI). WFP CARI (https://vamresources.manuals.wfp.org/docs/the-consolidated-approach-for-reporting-indicators-of-food-security-cari).",
        "weighting_rationale": "Three pillars weighted equally (1/3 each): IPC gives all three equal standing in convergence of evidence - no standard justifies weighting one pillar above another. Within FCS, Poor outweighs Borderline (70/30): Poor FCS (<21) is definitively inadequate diet; Borderline (21-35) is at-risk but not failing - qualitatively different states. rCSI and LCSI are summed as crisis-or-above totals (Phase 3 + Phase 4, Crisis + Emergency) before entering the formula - all crisis-level coping is treated equally regardless of severity tier.",
        "weight_note": "Composite scoring is not specified by IPC - IPC uses analyst-driven convergence, not a formula. Weights are a prototype design choice pending expert validation.",
        "indicators": {
            "poor_FCS": {
                "sheet": "FSL", "col": 59,
                "question": "What is the food consumption status of this household?",
                "response": "Poor",
                "effective_weight": round(0.70 / 3, 4),
                "weight_rationale": "0.70 within FCS pillar (1/3 of total): Poor FCS is definitively food insecure - dietary intake is unacceptably inadequate.",
            },
            "borderline_FCS": {
                "sheet": "FSL", "col": 58,
                "question": "What is the food consumption status of this household?",
                "response": "Borderline",
                "effective_weight": round(0.30 / 3, 4),
                "weight_rationale": "0.30 within FCS pillar (1/3 of total): Borderline FCS is at risk but diet not yet unacceptably poor - less severe than Poor.",
            },
            "rcsi_phase3_crisis": {
                "sheet": "FSL", "col": 63,
                "question": "What is the reduced coping strategy status of this household?",
                "response": "Phase 3 (Crisis, rCSI 10-18)",
                "effective_weight": round(1 / 3, 4),
                "weight_rationale": "Summed with rcsi_phase4 as crisis-or-above total; combined total carries 1/3 pillar weight. rCSI phase thresholds from WFP VAM (https://vamresources.manuals.wfp.org/docs/reduced-coping-strategies-index).",
            },
            "rcsi_phase4_emergency": {
                "sheet": "FSL", "col": 64,
                "question": "What is the reduced coping strategy status of this household?",
                "response": "Phase 4 (Emergency, rCSI >= 19)",
                "effective_weight": round(1 / 3, 4),
                "weight_rationale": "Summed with rcsi_phase3 as crisis-or-above total; combined total carries 1/3 pillar weight.",
            },
            "lcsi_crisis": {
                "sheet": "FSL", "col": 72,
                "question": "What is the livelihood coping status of this household?",
                "response": "Crisis coping strategies",
                "effective_weight": round(1 / 3, 4),
                "weight_rationale": "Summed with lcsi_emergency as crisis-or-above total; combined total carries 1/3 pillar weight.",
            },
            "lcsi_emergency": {
                "sheet": "FSL", "col": 73,
                "question": "What is the livelihood coping status of this household?",
                "response": "Emergencies coping strategies",
                "effective_weight": round(1 / 3, 4),
                "weight_rationale": "Summed with lcsi_crisis as crisis-or-above total; combined total carries 1/3 pillar weight.",
            },
        },
    },
    "WASH": {
        "formula": "(100 - improved_water) / 6 + insuff_water / 6 + open_defecation / 6 + unimproved_sanitation / 6 + (100 - soap_handwashing) / 3",
        "standard": "JMP 2023 methodology (https://washdata.org/topics/methods) - three independent pillars: Water, Sanitation, Hygiene. Sphere 2018 minimum 15L/person/day water quantity.",
        "weighting_rationale": "Three JMP pillars weighted equally (1/3 each): JMP explicitly monitors and reports Water, Sanitation, and Hygiene as separate equal pillars - no standard justifies weighting one above another. Within Water and Sanitation, two indicators each are split equally (1/6 per indicator). Hygiene has one indicator which carries the full 1/3 pillar weight. Consequence: hygiene (soap observation) carries twice the weight of any single water or sanitation indicator - a known limitation of having fewer hygiene indicators in this MSNA.",
        "weight_note": "JMP does not produce a composite WASH score - it reports each pillar independently. This formula is a prototype design choice pending expert validation.",
        "indicators": {
            "improved_water": {
                "sheet": "WASH", "col": 15,
                "question": "Access to improved drinking water",
                "response": "Improved water source",
                "pillar": "Water",
                "effective_weight": round(1 / 6, 4),
                "weight_rationale": "1/6: equal share of Water pillar (1/3) split between two water indicators.",
                "note": "Inverted: score uses (100 - value)",
            },
            "insuff_water": {
                "sheet": "WASH", "col": 27,
                "question": "Which of the following needs is your household able to meet with available water?",
                "response": "5. Not enough water to meet any of the above needs",
                "pillar": "Water",
                "effective_weight": round(1 / 6, 4),
                "weight_rationale": "1/6: equal share of Water pillar (1/3) split between two water indicators. Sphere 2018 minimum 15L/person/day.",
            },
            "open_defecation": {
                "sheet": "WASH", "col": 82,
                "question": "Type of sanitation facility used",
                "response": "None of the above, open defecation",
                "pillar": "Sanitation",
                "effective_weight": round(1 / 6, 4),
                "weight_rationale": "1/6: equal share of Sanitation pillar (1/3) split between two sanitation indicators.",
            },
            "unimproved_sanitation": {
                "sheet": "WASH", "col": 85,
                "question": "Access to improved sanitation",
                "response": "Unimproved sanitation type",
                "pillar": "Sanitation",
                "effective_weight": round(1 / 6, 4),
                "weight_rationale": "1/6: equal share of Sanitation pillar (1/3) split between two sanitation indicators.",
            },
            "soap_handwashing": {
                "sheet": "WASH", "col": 151,
                "question": "(Observe): Is soap available at the place for handwashing?",
                "response": "Yes",
                "pillar": "Hygiene",
                "effective_weight": round(1 / 3, 4),
                "weight_rationale": "1/3: sole indicator for Hygiene pillar carries full pillar weight. JMP Basic hygiene requires soap and water at handwashing station (https://washdata.org/monitoring/hygiene).",
                "note": "Inverted: score uses (100 - value)",
            },
        },
    },
    "Health": {
        "formula": "unable_to_access_care * 0.70 + facility_over_60min * 0.30",
        "weighting_rationale": "Unable to access care (0.70) outweighs distance (0.30): unable_to_access_care is an actual unmet need - the household had a health problem and could not obtain care. facility_over_60min is a proximity barrier - a household may still access care despite the distance. These are qualitatively different: realized deprivation vs potential barrier.",
        "weight_note": "No published standard specifies weights for a composite health access score. The 70/30 split reflects the outcome vs barrier distinction and is a prototype design choice pending expert validation.",
        "indicators": {
            "unable_to_access_care": {
                "sheet": "Health", "col": 12,
                "question": "Were households able to obtain health care?",
                "response": "No",
                "effective_weight": 0.70,
                "weight_rationale": "0.70: realized unmet need - household required care and could not access it.",
                "note": "Denominator is households that reported a health problem (col 6), not all households.",
            },
            "facility_over_60min": {
                "sheet": "Health", "col": 4,
                "question": "Travel time to nearest functional health facility",
                "response": "More than 60 minutes",
                "effective_weight": 0.30,
                "weight_rationale": "0.30: proximity barrier - travel time is a potential constraint, not a confirmed access failure.",
            },
        },
    },
    "Education": {
        "formula": "children_not_attending * 1.00",
        "weighting_rationale": "Single indicator - full weight by definition. No weighting decision required.",
        "weight_note": "Education sector has one indicator in this MSNA. Score equals the indicator value directly.",
        "indicators": {
            "children_not_attending": {
                "sheet": "Education", "col": 8,
                "question": "Did child attend formal school >= 4 days/week in past 6 months? (ages 5-18)",
                "response": "No",
                "effective_weight": 1.00,
                "weight_rationale": "Sole indicator for this sector.",
            },
        },
    },
    "Shelter": {
        "formula": "(100 - solid_shelter) * 0.25 + (makeshift + tent + no_shelter) * 0.25 + overcrowded * 0.25 + ((nfi_bedding + nfi_cooking) / 2) * 0.25",
        "standard": "Sphere Handbook 2018 (https://spherestandards.org/wp-content/uploads/Sphere-Handbook-2018-EN.pdf) - minimum 3.5m2/person floor area; core NFI items include sleeping materials and cooking equipment.",
        "weighting_rationale": "Four dimensions weighted equally (0.25 each): no published standard specifies weights for a composite shelter score. Equal weighting is the neutral prior in the absence of evidence. The four dimensions represent distinct Sphere-grounded concerns: (1) structure quality - what type of dwelling; (2) acute inadequacy - tents, makeshift, none; (3) space - Sphere 3.5m2 minimum; (4) core NFI - sleeping and cooking items. Known limitation: solid_shelter and acute_types are correlated (makeshift/tent/no shelter are a subset of not-solid-shelter), which partially double-counts structure inadequacy.",
        "weight_note": "Sphere specifies minimum thresholds per item independently - it does not define a composite score or weights. Equal weighting is a prototype design choice pending expert validation.",
        "indicators": {
            "solid_shelter": {
                "sheet": "Shelter", "col": 1,
                "question": "Type of shelter or dwelling",
                "response": "Solid / finished house",
                "effective_weight": 0.25,
                "weight_rationale": "0.25: equal weight across four shelter dimensions. Solid/finished houses only; apartments excluded.",
                "note": "Inverted: score uses (100 - value)",
            },
            "tent": {
                "sheet": "Shelter", "col": 6,
                "question": "Type of shelter or dwelling",
                "response": "Tent",
                "effective_weight": 0.25,
                "weight_rationale": "0.25 combined weight for acute shelter types (makeshift + tent + no_shelter) - summed before applying weight.",
            },
            "makeshift": {
                "sheet": "Shelter", "col": 7,
                "question": "Type of shelter or dwelling",
                "response": "Makeshift shelter",
                "effective_weight": 0.25,
                "weight_rationale": "0.25 combined weight for acute shelter types - summed with tent and no_shelter before applying weight.",
            },
            "no_shelter": {
                "sheet": "Shelter", "col": 10,
                "question": "Type of shelter or dwelling",
                "response": "No shelter (sleeping in the open)",
                "effective_weight": 0.25,
                "weight_rationale": "0.25 combined weight for acute shelter types - summed with makeshift and tent before applying weight.",
            },
            "overcrowded": {
                "sheet": "Shelter", "col": 20,
                "question": "What issues are present in your current shelter?",
                "response": "9. Overcrowding (less than 3.5 m2 per household member)",
                "effective_weight": 0.25,
                "weight_rationale": "0.25: equal weight. Sphere 2018 minimum 3.5m2/person in emergency settings.",
            },
            "nfi_bedding": {
                "sheet": "Shelter", "col": 38,
                "question": "Please explain why your household cannot sleep / the issues",
                "response": "1. Insufficient essential household items for sleeping (bedding, mattresses/mats, bednets)",
                "effective_weight": 0.125,
                "weight_rationale": "0.125: half of 0.25 NFI dimension weight, averaged with nfi_cooking. Sphere 2018 core NFI item.",
            },
            "nfi_cooking": {
                "sheet": "Shelter", "col": 51,
                "question": "Please explain why your household cannot cook / the issues",
                "response": "1. Insufficient essential household items for cooking (utensils, kitchen sets, eating sets)",
                "effective_weight": 0.125,
                "weight_rationale": "0.125: half of 0.25 NFI dimension weight, averaged with nfi_bedding. Sphere 2018 core NFI item.",
            },
        },
    },
    "Protection": {
        "formula": "(movement_restrictions + missing_civil_docs + psychological_distress + unaccompanied_children) / 4",
        "weighting_rationale": "Four indicators weighted equally (0.25 each): movement restrictions, civil documentation, psychological distress, and unaccompanied children measure independent protection dimensions with no published cross-dimension weighting. Equal weighting is the neutral prior.",
        "weight_note": "No published standard specifies weights for a composite multi-sector protection score. Equal weighting is a prototype design choice pending expert validation.",
        "indicators": {
            "movement_restrictions": {
                "sheet": "General_Protection_CP_GBV", "col": 10,
                "question": "In the last 3 months, have any members of your household faced restrictions when trying to move?",
                "response": "Yes",
                "effective_weight": 0.25,
                "weight_rationale": "0.25: equal weight across four independent protection dimensions.",
            },
            "missing_civil_docs": {
                "sheet": "General_Protection_CP_GBV", "col": 24,
                "question": "Are any household members missing their civil documentation (passport, birth certificate, national ID, etc.)?",
                "response": "Not missing",
                "effective_weight": 0.25,
                "weight_rationale": "0.25: equal weight. Derived: score uses (100 - col 24) to convert % not-missing to % missing at least one document.",
                "note": "Inverted: col 24 gives % with no missing docs; score uses (100 - value) to get % missing at least one document.",
            },
            "psychological_distress": {
                "sheet": "General_Protection_CP_GBV", "col": 55,
                "question": "Have any members of your household experienced any signs of psychological distress?",
                "response": "Yes",
                "effective_weight": 0.25,
                "weight_rationale": "0.25: equal weight across four independent protection dimensions.",
            },
            "unaccompanied_children": {
                "sheet": "General_Protection_CP_GBV", "col": 76,
                "question": "Are there children under 18 now living in this household without both of their parents?",
                "response": "Yes",
                "effective_weight": 0.25,
                "weight_rationale": "0.25: equal weight across four independent protection dimensions.",
            },
        },
    },
}

# ─── Manual pCode overrides ───────────────────────────────────────────────────
# Used when automatic name matching is ambiguous. Each entry documents its
# reasoning so the assignment can be reviewed and confirmed later.

MANUAL_PCODE_OVERRIDES: dict = {
    "ar rahad": {
        "p_code":       "SD13030",
        "match_method": "provisional_manual",
        "match_note": (
            "Two OCHA COD localities share this name: SD13030 (North Kordofan) and SD12084 (Gedaref). "
            "MSNA Demography sheet records 62.4% urban for this locality. "
            "SD13030 is a major city and railway hub - the second largest locality in North Kordofan "
            "(~26,000 population, 2012 census) - consistent with a 62.4% urban profile. "
            "SD12084 is the Rahad agricultural irrigation scheme area in Gedaref, expected to be "
            "predominantly rural (cf. Gedaref siblings: Al Butanah 0%, Al Fashaga 31%, "
            "Galabat Ash-Shargiah 30% urban). "
            "Provisional assignment to SD13030 pending confirmation from IOM Sudan DTM."
        ),
    },
}

# ─── Severity thresholds ──────────────────────────────────────────────────────
# Source: Severity Scoring Logic.md §1

THRESHOLDS = [
    (80, "Catastrophic"),
    (65, "Extreme"),
    (45, "Severe"),
    (25, "Stressed"),
    (0,  "Minimal"),
]

def score_to_label(score: float | None) -> str | None:
    if score is None:
        return None
    for min_score, label in THRESHOLDS:
        if score >= min_score:
            return label
    return "Minimal"


# ─── Column definitions ───────────────────────────────────────────────────────
# Format: indicator_key -> (sheet_name, col_index, r1_fragment, r2_fragment)
#
# r1_fragment and r2_fragment are case-insensitive substrings used to verify
# that the column at col_index actually contains what we expect. The script
# aborts if any check fails.

COLUMN_DEFS = {
    # Food Security - FSL sheet
    "fsl_poor_fcs":           ("FSL",       59, "food consumption status",     "Poor"),
    "fsl_borderline_fcs":     ("FSL",       58, "food consumption status",     "Borderline"),
    "fsl_mod_food_insecure":  ("FSL",       79, "food security status",        "Moderately Food Insecure"),
    "fsl_sev_food_insecure":  ("FSL",       80, "food security status",        "Severely Food Insecure"),
    # WASH
    "wash_improved_water":    ("WASH",      15, "improved drinking water",     "Improved water source"),
    "wash_open_defecation":   ("WASH",      82, "sanitation facility",         "open defecation"),
    "wash_unimp_sanitation":  ("WASH",      85, "improved sanitation",         "Unimproved sanitation"),
    # Health
    # Col 12 = "No" on "were households able to obtain health care" → unable to access
    # Denominator: households that reported a health problem (Col 6), not all households
    "health_unable_access":   ("Health",    12, "able to obtain health care",  "No"),
    "health_over_60min":      ("Health",     4, "travel to nearest",           "More than 60 minutes"),
    # Education
    # Col 8 = "No" on "did child attend formal school" → not attending
    "edu_not_attending":      ("Education",  8, "attend formal school",        "No"),
    # Shelter
    "shelter_solid":          ("Shelter",    1, "type of shelter",             "Solid"),
    "shelter_tent":           ("Shelter",    6, "type of shelter",             "Tent"),
    "shelter_makeshift":      ("Shelter",    7, "type of shelter",             "Makeshift"),
    "shelter_none":           ("Shelter",   10, "type of shelter",             "No shelter"),
    # FSL - rCSI phase breakdown (IPC three-pillar)
    "fsl_rcsi_phase3":     ("FSL",      63, "reduce coping stratergy",               "Phase 3"),
    "fsl_rcsi_phase4":     ("FSL",      64, "reduce coping stratergy",               "Phase 4"),
    # FSL - LCSI livelihood coping (IPC three-pillar)
    "fsl_lcsi_crisis":     ("FSL",      72, "livelihood coping status",              "Crisis coping"),
    "fsl_lcsi_emergency":  ("FSL",      73, "livelihood coping status",              "Emergencies coping"),
    # WASH - water quantity (Sphere 15L/person/day minimum)
    "wash_insuff_water":   ("WASH",     27, "needs is your household able to meet",  "Not enough water"),
    # WASH - soap at handwashing station (JMP Basic hygiene)
    "wash_soap_hwashing":  ("WASH",    151, "soap available at the place",           "Yes"),
    # Shelter - overcrowding (Sphere 3.5m2/person)
    "shelter_overcrowded": ("Shelter",  20, "what issues, if any, are present",      "Overcrowding"),
    # Shelter - NFI gaps (Sphere core relief items)
    "shelter_nfi_bedding": ("Shelter",  38, "explain why your household can't sleep","Insufficient essential household items for sleeping"),
    "shelter_nfi_cooking": ("Shelter",  51, "explain why your household can't cook", "Insufficient essential household items for cooking"),
    # Protection - General_Protection_CP_GBV sheet
    # prot_not_missing_docs is the "Not missing" % - inverted in score_protection to get missing%
    "prot_movement_restricted": ("General_Protection_CP_GBV", 10, "faced restrictions",                    "Yes"),
    "prot_not_missing_docs":    ("General_Protection_CP_GBV", 24, "missing their civil documentation",      "Not missing"),
    "prot_psych_distress":      ("General_Protection_CP_GBV", 55, "psychological distress",                 "Yes"),
    "prot_unaccomp_children":   ("General_Protection_CP_GBV", 76, "children under 18",                     "Yes"),
}


# ─── Scoring formulas ─────────────────────────────────────────────────────────
# Source: Severity Scoring Logic.md §2

def score_fsl(row: dict) -> tuple[float, dict]:
    poor       = row["fsl_poor_fcs"]
    borderline = row["fsl_borderline_fcs"]
    rcsi_p3    = row["fsl_rcsi_phase3"]
    rcsi_p4    = row["fsl_rcsi_phase4"]
    lcsi_cr    = row["fsl_lcsi_crisis"]
    lcsi_em    = row["fsl_lcsi_emergency"]

    fcs_score  = (poor or 0) * 0.70 + (borderline or 0) * 0.30
    rcsi_score = (rcsi_p3 or 0) + (rcsi_p4 or 0)
    lcsi_score = (lcsi_cr or 0) + (lcsi_em or 0)
    score = (fcs_score + rcsi_score + lcsi_score) / 3

    inputs = {
        "poor_FCS":              poor,
        "borderline_FCS":        borderline,
        "rcsi_phase3_crisis":    rcsi_p3,
        "rcsi_phase4_emergency": rcsi_p4,
        "lcsi_crisis":           lcsi_cr,
        "lcsi_emergency":        lcsi_em,
    }
    return round(score, 2), inputs


def score_wash(row: dict) -> tuple[float, dict]:
    improved  = row["wash_improved_water"]
    open_def  = row["wash_open_defecation"]
    unimp_san = row["wash_unimp_sanitation"]
    insuff    = row["wash_insuff_water"]
    soap      = row["wash_soap_hwashing"]

    inputs = {
        "improved_water":        improved,
        "insuff_water":          insuff,
        "open_defecation":       open_def,
        "unimproved_sanitation": unimp_san,
        "soap_handwashing":      soap,
    }
    # JMP three pillars equal (1/3 each): Water = quality/6 + quantity/6,
    # Sanitation = OD/6 + unimproved/6, Hygiene = soap/3
    score = (
        (100 - (improved or 0)) / 6
        + (insuff         or 0) / 6
        + (open_def       or 0) / 6
        + (unimp_san      or 0) / 6
        + (100 - (soap   or 0)) / 3
    )
    return round(score, 2), inputs


def score_health(row: dict) -> tuple[float, dict]:
    inputs = {
        "unable_to_access_care": row["health_unable_access"],
        "facility_over_60min":   row["health_over_60min"],
    }
    score = (
        inputs["unable_to_access_care"] * 0.70
        + inputs["facility_over_60min"] * 0.30
    )
    return round(score, 2), inputs


def score_education(row: dict) -> tuple[float, dict]:
    inputs = {"children_not_attending": row["edu_not_attending"]}
    score = inputs["children_not_attending"]
    return round(score, 2), inputs


def score_shelter(row: dict) -> tuple[float, dict]:
    solid       = row["shelter_solid"]
    makeshift   = row["shelter_makeshift"]
    tent        = row["shelter_tent"]
    no_shelt    = row["shelter_none"]
    overcrowded = row["shelter_overcrowded"]
    nfi_bedding = row["shelter_nfi_bedding"]
    nfi_cooking = row["shelter_nfi_cooking"]

    acute_types = (makeshift or 0) + (tent or 0) + (no_shelt or 0)
    nfi_avg     = ((nfi_bedding or 0) + (nfi_cooking or 0)) / 2.0

    inputs = {
        "solid_shelter":       solid,
        "makeshift":           makeshift,
        "tent":                tent,
        "no_shelter":          no_shelt,
        "overcrowded":         overcrowded,
        "nfi_lacking_bedding": nfi_bedding,
        "nfi_lacking_cooking": nfi_cooking,
    }
    score = (
        (100 - (solid or 0)) * 0.25
        + acute_types        * 0.25
        + (overcrowded or 0) * 0.25
        + nfi_avg            * 0.25
    )
    return round(score, 2), inputs


def score_protection(row: dict) -> tuple[float, dict]:
    movement    = row["prot_movement_restricted"]
    not_missing = row["prot_not_missing_docs"]
    distress    = row["prot_psych_distress"]
    unaccomp    = row["prot_unaccomp_children"]
    # Invert "Not missing" to get % of households missing at least one document
    missing_docs = (100.0 - not_missing) if not_missing is not None else None
    inputs = {
        "movement_restrictions":  movement,
        "missing_civil_docs":     missing_docs,
        "psychological_distress": distress,
        "unaccompanied_children": unaccomp,
    }
    score = (
        (movement     or 0) * 0.25
        + (missing_docs or 0) * 0.25
        + (distress     or 0) * 0.25
        + (unaccomp     or 0) * 0.25
    )
    return round(score, 2), inputs


SECTOR_SCORERS: dict = {
    "FSL":        score_fsl,
    "WASH":       score_wash,
    "Health":     score_health,
    "Education":  score_education,
    "Shelter":    score_shelter,
    "Protection": score_protection,
}


# ─── xlsx parsing ─────────────────────────────────────────────────────────────

def parse_percent(val) -> float | None:
    if val is None:
        return None
    s = str(val).strip().rstrip("%")
    try:
        return float(s)
    except ValueError:
        return None


def _ffill(row: tuple) -> list:
    """Forward-fill None values across a row (for merged header cells)."""
    out = list(row)
    last = None
    for i, v in enumerate(out):
        if v is not None:
            last = v
        else:
            out[i] = last
    return out


def load_all_raw_data(xlsx_path: Path) -> dict[str, dict[str, dict]]:
    """
    Returns {locality_name: {sheet_name: {col_label: value}}} for all columns
    in every sheet referenced by COLUMN_DEFS. Column 0 (locality name) is
    excluded. col_label is built from the 2-row merged header
    ("{r1_ffilled} | {r2}"); duplicate labels within a sheet get a _2, _3 …
    suffix. Values are parsed as percentages (float) where possible, else None.
    """
    target_sheets = list(dict.fromkeys(v[0] for v in COLUMN_DEFS.values()))
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    result: dict[str, dict] = {}

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found (raw)", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        r1_filled = _ffill(rows[0])
        r2 = list(rows[1])

        # Build column labels from 2-row merged header
        raw_keys: list[str] = []
        for i, r1_val in enumerate(r1_filled):
            r1s = str(r1_val).strip() if r1_val is not None else ""
            r2s = str(r2[i]).strip() if i < len(r2) and r2[i] is not None else ""
            if r1s and r2s:
                raw_keys.append(f"{r1s} | {r2s}")
            elif r1s:
                raw_keys.append(r1s)
            else:
                raw_keys.append(f"col_{i}")

        # Deduplicate: append _2, _3, ... to later occurrences of the same label
        seen: dict[str, int] = {}
        col_keys: list[str] = []
        for k in raw_keys:
            if k not in seen:
                seen[k] = 1
                col_keys.append(k)
            else:
                seen[k] += 1
                col_keys.append(f"{k}_{seen[k]}")

        for row in rows[2:]:
            locality = str(row[0]).strip() if row[0] else None
            if not locality or locality.lower() in ("none", "locality", "population group"):
                continue
            sheet_data: dict[str, float | None] = {}
            for i in range(1, len(col_keys)):  # col 0 is locality name, skip
                val = row[i] if i < len(row) else None
                sheet_data[col_keys[i]] = parse_percent(val)
            result.setdefault(locality, {})[sheet_name] = sheet_data

    wb.close()
    return result


def load_all_indicators(xlsx_path: Path) -> dict[str, dict]:
    """
    Returns {locality_name: {indicator_key: float_value}} for all localities
    and all columns defined in COLUMN_DEFS.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Group columns by sheet
    by_sheet: dict[str, dict] = {}
    for key, (sheet, col_idx, _, _) in COLUMN_DEFS.items():
        by_sheet.setdefault(sheet, {})[key] = col_idx

    merged: dict[str, dict] = {}
    for sheet_name, cols in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: sheet '{sheet_name}' not found", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[2:]:  # skip 2-row header
            locality = str(row[0]).strip() if row[0] else None
            if not locality or locality.lower() in ("none", "locality", "population group"):
                continue
            entry = merged.setdefault(locality, {})
            for key, col_idx in cols.items():
                entry[key] = parse_percent(row[col_idx] if col_idx < len(row) else None)

    wb.close()
    return merged


# ─── Column verification ──────────────────────────────────────────────────────

def verify_columns(xlsx_path: Path, silent: bool = False) -> bool:
    """
    Checks every entry in COLUMN_DEFS against the actual xlsx headers.
    Prints a verification table. Returns True if all checks pass.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if not silent:
        print(f"\n{'INDICATOR':<28} {'SHEET':<12} {'COL':>4}  {'STATUS':<6}  HEADERS")
        print("─" * 120)

    all_pass = True
    for indicator, (sheet, col_idx, exp_r1, exp_r2) in COLUMN_DEFS.items():
        if sheet not in wb.sheetnames:
            if not silent:
                print(f"{indicator:<28} {sheet:<12} {col_idx:>4}  FAIL    sheet not found in workbook")
            all_pass = False
            continue

        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
        r1_filled = _ffill(rows[0])

        actual_r1 = str(r1_filled[col_idx]).strip() if col_idx < len(r1_filled) and r1_filled[col_idx] else ""
        actual_r2 = str(rows[1][col_idx]).strip() if col_idx < len(rows[1]) and rows[1][col_idx] else ""

        r1_ok = exp_r1.lower() in actual_r1.lower()
        r2_ok = exp_r2.lower() in actual_r2.lower()
        ok = r1_ok and r2_ok

        if not ok:
            all_pass = False

        if not silent:
            status = "PASS" if ok else "FAIL"
            r1_disp = actual_r1[:55].replace("\n", " ") + ("…" if len(actual_r1) > 55 else "")
            r2_disp = actual_r2[:30].replace("\n", " ")
            r1_flag = "" if r1_ok else " ← r1 mismatch"
            r2_flag = "" if r2_ok else " ← r2 mismatch"
            print(f"{indicator:<28} {sheet:<12} {col_idx:>4}  {status:<6}  r2={r2_disp!r}{r2_flag}")
            if not ok:
                print(f"  {'':28} {'':12}       expected r1 fragment: {exp_r1!r}{r1_flag}")
                print(f"  {'':28} {'':12}       actual   r1: {r1_disp!r}")

    wb.close()
    if not silent:
        print()
    return all_pass


# ─── pCode lookup via CLEAR API ───────────────────────────────────────────────

def normalize_name(name: str) -> str:
    s = name.lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def fetch_api_localities() -> dict[str, dict]:
    """Returns {normalized_name: {id, name, p_code}} via the public CLEAR GraphQL API."""
    query = '{ locations(level: 2) { id name pCode parent { name } } }'
    payload = json.dumps({"query": query}).encode()
    req = urllib.request.Request(
        CLEAR_API_URL,
        data=payload,
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        data = json.loads(resp.read())

    errors = data.get("errors")
    if errors:
        print(f"ERROR: CLEAR API returned errors: {errors}", file=sys.stderr)
        sys.exit(1)

    locations = data["data"]["locations"]
    return {
        normalize_name(loc["name"]): {
            "id":     loc["id"],
            "name":   loc["name"],
            "p_code": loc["pCode"],
        }
        for loc in locations
    }


def match_locality(msna_name: str, api_locs: dict) -> dict | None:
    norm = normalize_name(msna_name)

    # 0. Manual overrides - documented provisional assignments for ambiguous names
    if norm in MANUAL_PCODE_OVERRIDES:
        override = MANUAL_PCODE_OVERRIDES[norm]
        # Look up the location record by pCode so we get the id too
        loc_rec = next((v for v in api_locs.values() if v["p_code"] == override["p_code"]), None)
        if loc_rec:
            return {
                **loc_rec,
                "match_method": override["match_method"],
                "match_note":   override["match_note"],
            }

    # 1. Exact normalized match
    if norm in api_locs:
        return {**api_locs[norm], "match_method": "exact"}

    # 2. Strip leading Arabic definite article prefixes from the MSNA name
    for prefix in ("al ", "el ", "ad ", "az ", "as ", "at ", "ar "):
        if norm.startswith(prefix):
            candidate = norm[len(prefix):]
            if candidate in api_locs:
                return {**api_locs[candidate], "match_method": "prefix_stripped"}

    # 3. Strip prefix from the API name side
    for api_norm, api_rec in api_locs.items():
        for prefix in ("al ", "el ", "ad ", "az ", "as ", "at ", "ar "):
            if api_norm.startswith(prefix) and api_norm[len(prefix):] == norm:
                return {**api_rec, "match_method": "db_prefix_stripped"}

    return None


# ─── Processing ───────────────────────────────────────────────────────────────

def process(xlsx_path: Path, api_locs: dict | None) -> dict:
    raw_data = load_all_indicators(xlsx_path)
    all_raw = load_all_raw_data(xlsx_path)
    localities_out = {}
    unmatched = []

    sector_scorers = SECTOR_SCORERS

    for locality_name in sorted(raw_data):
        indicators = raw_data[locality_name]
        entry: dict = {}

        # pCode matching
        if api_locs is not None:
            match = match_locality(locality_name, api_locs)
            if match:
                entry["pcode"]        = match["p_code"]
                entry["location_id"]  = match["id"]
                entry["match_method"] = match["match_method"]
                if "match_note" in match:
                    entry["match_note"] = match["match_note"]
            else:
                entry["pcode"]        = None
                entry["location_id"]  = None
                entry["match_method"] = "unmatched"
                unmatched.append(locality_name)

        # Sector scores
        sectors = {}
        for sector, scorer in sector_scorers.items():
            if scorer is None:
                sectors[sector] = {
                    "note":   "formula_missing",
                    "score":  None,
                    "label":  None,
                    "inputs": {},
                }
                continue
            try:
                score, inputs = scorer(indicators)
                sectors[sector] = {
                    "score":  score,
                    "label":  score_to_label(score),
                    "inputs": inputs,
                }
            except Exception as exc:
                sectors[sector] = {
                    "error":  str(exc),
                    "score":  None,
                    "label":  None,
                    "inputs": {},
                }

        entry["sectors"] = sectors

        if locality_name in all_raw:
            entry["raw"] = all_raw[locality_name]

        localities_out[locality_name] = entry

    matched = sum(1 for v in localities_out.values() if v.get("pcode") is not None)

    return {
        "meta": {
            "generated_at":        datetime.now(timezone.utc).isoformat(),
            "source_file":         xlsx_path.name,
            "source_title":        "IOM/DTM - Multi-Sector Needs Assessment - Sudan 2025",
            "data_collection":     "07 August 2025 to 31 August 2025",
            "led_by":              "IOM DTM, with OCHA, REACH, ICCG, and AAWG",
            "coverage":            "188 localities across all 18 Sudan states",
            "scoring_version":     "prototype-v3",
            "formula_source":      FORMULA_SOURCE,
            "note":                "Thresholds and formulas are prototype-stage and unvalidated.",
            "formulas":            FORMULA_DOCS,
        },
        "summary": {
            "total_localities": len(localities_out),
            "matched_to_pcode":  matched if api_locs is not None else "skipped",
            "unmatched_count":  len(unmatched),
        },
        "unmatched_localities": unmatched,
        "localities":           localities_out,
    }


# ─── Entry point ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="MSNA severity scoring script")
    parser.add_argument("--verify",     action="store_true", help="Verify column mappings and exit")
    parser.add_argument("--no-db",      action="store_true", help="Skip pCode DB lookup (offline mode)")
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    args = parser.parse_args()

    if args.verify:
        ok = verify_columns(XLSX_PATH)
        sys.exit(0 if ok else 1)

    # Always verify before processing
    print("Verifying column mappings...")
    ok = verify_columns(XLSX_PATH)
    if not ok:
        print("Column verification FAILED - aborting. Fix COLUMN_DEFS before processing.", file=sys.stderr)
        sys.exit(1)
    print("All columns verified.\n")

    api_locs = None
    if not args.no_db:
        print("Fetching localities from CLEAR API...")
        try:
            api_locs = fetch_api_localities()
            print(f"  Loaded {len(api_locs)} Admin2 localities\n")
        except Exception as exc:
            print(f"ERROR: Could not reach CLEAR API: {exc}", file=sys.stderr)
            print("Run with --no-db to skip pCode matching.", file=sys.stderr)
            sys.exit(1)

    print("Processing MSNA data...")
    result = process(XLSX_PATH, api_locs)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime("%Y-%m-%d")
    out_path = args.output_dir / f"msna_{date_str}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f"Output written to: {out_path}")
    print(f"  Localities:   {result['summary']['total_localities']}")
    if api_locs is not None:
        print(f"  Matched:      {result['summary']['matched_to_pcode']}")
        print(f"  Unmatched:    {result['summary']['unmatched_count']}")
        if result["unmatched_localities"]:
            print("\nUnmatched localities (need manual pCode assignment):")
            for name in result["unmatched_localities"]:
                print(f"  - {name}")


if __name__ == "__main__":
    main()
